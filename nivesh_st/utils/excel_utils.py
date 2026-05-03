"""
Comprehensive Excel export utility — multi-sheet workbooks with full fundamentals.
Used by Alpha Hunter, Screener, Watchlist, Compare.
"""
from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime

# ── Colour palette ─────────────────────────────────────────────────────────
C_GOLD    = "F59E0B"
C_GREEN   = "10D98D"
C_RED     = "FF4757"
C_NAVY    = "1E3A8A"
C_DARK    = "111827"
C_HEADER  = "0D111C"
C_WHITE   = "F1F5F9"
C_GREY    = "8DA4BF"

FILL_HEADER = PatternFill("solid", fgColor=C_HEADER)
FILL_GOLD   = PatternFill("solid", fgColor=C_GOLD)
FILL_DARK   = PatternFill("solid", fgColor=C_DARK)
FILL_NAVY   = PatternFill("solid", fgColor=C_NAVY)
FILL_SUBHDR = PatternFill("solid", fgColor="1E2A3A")

FONT_HEADER = Font(color=C_WHITE, bold=True, size=11, name="Calibri")
FONT_GOLD   = Font(color=C_GOLD, bold=True, size=10, name="Calibri")
FONT_WHITE  = Font(color=C_WHITE, size=10, name="Calibri")
FONT_GREY   = Font(color=C_GREY, size=9, name="Calibri")
FONT_TITLE  = Font(color=C_GOLD, bold=True, size=14, name="Calibri")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

THIN_BORDER = Border(
    left=Side(style="thin", color="1E2A3A"),
    right=Side(style="thin", color="1E2A3A"),
    bottom=Side(style="thin", color="1E2A3A"),
)

# ── Column definitions for each sheet ──────────────────────────────────────

SUMMARY_COLS = [
    ("Symbol",           "NSE Symbol",                         12),
    ("Company",          "Company Name",                       28),
    ("Sector",           "Sector",                             14),
    ("Price (₹)",        "Last Traded Price",                  12),
    ("Day Change %",     "Day Change %",                       12),
    ("Composite Score",  "Score /120 (Fund+Macro+Tech)",       16),
    ("Macro Boost",      "Live Macro Sector Impact",           14),
    ("Tech Signal",      "Technical Signal",                   14),
    ("Expected Return %","Prob-Weighted 12M Return Est.",      16),
    ("Analyst Target",   "Analyst Consensus Target (₹)",       16),
    ("Analyst Upside %", "Upside to Analyst Target",           14),
]

FUNDAMENTALS_COLS = [
    ("Symbol",               "NSE Symbol",                   10),
    ("Company",              "Company Name",                  28),
    ("Sector",               "Sector",                        14),
    ("Price (₹)",            "Last Traded Price (₹)",          12),
    ("Market Cap (₹Cr)",     "Market Capitalisation (₹Cr)",   18),
    ("P/E (TTM)",            "Price/Earnings Trailing 12M",   14),
    ("P/E (Forward)",        "Forward Price/Earnings",         14),
    ("Sector PE",            "Median PE of Sector",           12),
    ("PE vs Sector",         "PE Premium/Discount to Sector", 16),
    ("P/B Ratio",            "Price to Book Value",            12),
    ("P/S Ratio",            "Price to Sales (TTM)",           12),
    ("EV/EBITDA",            "Enterprise Value / EBITDA",      12),
    ("EV/Revenue",           "Enterprise Value / Revenue",     12),
    ("EPS (₹)",              "Earnings Per Share (₹)",          10),
    ("PEG Ratio",            "PE/Growth (< 1 = undervalued)",  12),
    ("Graham Number (₹)",    "Intrinsic Value = √(22.5×EPS×BV)",  16),
    ("Price vs Graham",      "% Premium/Discount to Graham",   16),
    ("ROE %",                "Return on Equity %",             10),
    ("ROA %",                "Return on Assets %",             10),
    ("ROCE %",               "Return on Capital Employed %",   10),
    ("Net Margin %",         "Net Profit Margin %",            12),
    ("Operating Margin %",   "Operating/EBIT Margin %",        14),
    ("Gross Margin %",       "Gross Profit Margin %",           12),
    ("Revenue Growth %",     "YoY Revenue Growth %",           14),
    ("EPS Growth %",         "YoY EPS / Earnings Growth %",    14),
    ("Debt/Equity",          "Total Debt / Equity",             10),
    ("Current Ratio",        "Current Assets / Current Liab.", 14),
    ("Interest Coverage",    "EBITDA / Interest Expense",       16),
    ("FCF (₹Cr)",            "Free Cash Flow (₹Cr)",             12),
    ("Payout Ratio %",       "Dividend Payout %",               12),
    ("Dividend Yield %",     "Annual Dividend Yield %",         14),
    ("Promoter Holding %",   "Insider/Promoter Holding %",      16),
    ("FII Holding %",        "Foreign Institutional Holding %", 16),
    ("Institutional Hold %", "Total Institutional Holding %",   18),
    ("52W High (₹)",         "52-Week High Price (₹)",           12),
    ("52W Low (₹)",          "52-Week Low Price (₹)",            12),
    ("52W Position %",       "Position in 52W Range (0=low,100=high)", 18),
    ("Beta",                 "Beta vs Nifty 50",                8),
    ("Analyst Target (₹)",   "Analyst Consensus Target",        16),
    ("Analyst Upside %",     "Upside to Analyst Target %",      16),
    ("Recommendation",       "Analyst Consensus Rec.",          14),
]

TECHNICALS_COLS = [
    ("Symbol",         "NSE Symbol",              10),
    ("Company",        "Company Name",             28),
    ("Price (₹)",      "Last Traded Price (₹)",     12),
    ("RSI (14)",       "RSI — <30 oversold, >70 overbought", 14),
    ("MACD",           "MACD Line",                12),
    ("MACD Signal",    "MACD Signal Line",          12),
    ("SMA 20 (₹)",     "20-Day Simple Moving Average", 14),
    ("SMA 50 (₹)",     "50-Day Simple Moving Average", 14),
    ("SMA 200 (₹)",    "200-Day Simple Moving Average",16),
    ("BB Upper (₹)",   "Bollinger Band Upper",      14),
    ("BB Lower (₹)",   "Bollinger Band Lower",      14),
    ("ATR (14) (₹)",   "Average True Range (₹)",    14),
    ("Stop 1.5×ATR",   "Tight Stop Loss (1.5×ATR)", 14),
    ("Stop 2.5×ATR",   "Normal Stop Loss (2.5×ATR)",14),
    ("Volume",         "Today's Volume",            14),
    ("Avg Volume 3M",  "3-Month Average Volume",    16),
    ("Vol vs Avg",     "Volume/Avg Volume Ratio",   14),
    ("Tech Signal",    "Overall Technical Signal",  14),
    ("Momentum %",     "10D vs Prior 10D Momentum", 12),
]

def _style_header_row(ws, row_idx: int, ncols: int, fill=None):
    f = fill or FILL_HEADER
    for col in range(1, ncols+1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = f
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

def _style_data_rows(ws, start_row: int, end_row: int, ncols: int):
    for row in range(start_row, end_row+1):
        fill = PatternFill("solid", fgColor=C_DARK if row%2==0 else "0D111C")
        for col in range(1, ncols+1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = Font(color=C_WHITE, size=10, name="Calibri")
            cell.alignment = ALIGN_LEFT
            cell.border = THIN_BORDER

def _set_col_widths(ws, col_defs):
    for i, (_, _, width) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def _add_title_row(ws, title: str, ncols: int):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = FILL_GOLD
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER

def _add_meta_sheet(wb, universe_size: int, macro_regime: dict = None):
    ws = wb.create_sheet("ℹ️ About")
    ws.sheet_view.showGridLines = False
    meta = [
        ("NIVESH Investment Platform", ""),
        ("", ""),
        ("Download Date/Time", datetime.now().strftime("%d %b %Y, %H:%M IST")),
        ("Universe", f"{universe_size} NSE stocks (Nifty50+Next50+Midcap+Smallcap+Extra)"),
        ("Score Formula", "Composite = Valuation(25) + Quality(30) + Growth(25) + Safety(20) + Technical(10) + LiveMacro(±20)"),
        ("", ""),
        ("PEG Ratio", "< 0.5 = Very Cheap | 0.5–1.0 = Cheap | 1.0–1.5 = Fair | > 2.0 = Expensive"),
        ("Graham Number", "sqrt(22.5 × EPS × Book Value) — Benjamin Graham's intrinsic value estimate"),
        ("Interest Coverage", "EBITDA / Interest Expense — >3x safe, <1.5x risky"),
        ("Macro Boost", "Live sector impact: crude, USD/INR, VIX, US market, rate regime"),
        ("RSI", "< 30 oversold (buy signal) | > 70 overbought (sell signal)"),
        ("Tech Signal", "Bullish = above SMA20+50+200 | Bearish = below all"),
        ("52W Position %", "0% = at 52W low, 100% = at 52W high"),
        ("", ""),
    ]
    if macro_regime:
        meta.append(("LIVE MACRO AT DOWNLOAD TIME", ""))
        for k, v in macro_regime.items():
            if not isinstance(v, bool) and k not in ["crude_high","crude_low","inr_weak","inr_strong","high_vix","us_bull","us_bear","rate_cut","rate_hike","mkt_up","mkt_down"]:
                meta.append((f"  {k.replace('_',' ').title()}", str(round(v,2) if isinstance(v,float) else v)))
    for i, (k, v) in enumerate(meta, 1):
        ws.cell(row=i, column=1, value=k).font = Font(color=C_GOLD, bold=True, size=10) if k and not k.startswith("  ") else Font(color=C_WHITE, size=10)
        ws.cell(row=i, column=2, value=v).font = Font(color=C_WHITE, size=10)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

def build_excel(summary_df: pd.DataFrame,
                fundamentals_df: pd.DataFrame = None,
                technicals_df: pd.DataFrame = None,
                sheet_name: str = "NIVESH Data",
                universe_size: int = 545,
                macro_regime: dict = None) -> bytes:
    """
    Build a styled multi-sheet Excel workbook.
    Sheet 1: Summary | Sheet 2: Full Fundamentals | Sheet 3: Technicals | Sheet 4: About
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Sheet 1 — Summary
        summary_df.to_excel(writer, index=False, sheet_name="📊 Summary")
        # Sheet 2 — Fundamentals
        if fundamentals_df is not None and not fundamentals_df.empty:
            fundamentals_df.to_excel(writer, index=False, sheet_name="💰 Fundamentals")
        # Sheet 3 — Technicals
        if technicals_df is not None and not technicals_df.empty:
            technicals_df.to_excel(writer, index=False, sheet_name="📈 Technicals")

    # Re-open for styling
    wb = load_workbook(buf)

    for sheet_title, ws in wb.worksheets:
        pass  # iterate below

    # Style each sheet
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        max_col = ws.max_column
        max_row = ws.max_row

        # Add title row at top
        title_map = {
            "📊 Summary":       f"NIVESH — {sheet_name} | {datetime.now().strftime('%d %b %Y')}",
            "💰 Fundamentals":  f"Full Fundamental Metrics | {universe_size} stocks | {datetime.now().strftime('%d %b %Y')}",
            "📈 Technicals":    f"Technical Indicators | {datetime.now().strftime('%d %b %Y')}",
        }
        if ws.title in title_map:
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=1, value=title_map[ws.title])
            title_cell.fill = PatternFill("solid", fgColor=C_NAVY)
            title_cell.font = Font(color=C_GOLD, bold=True, size=13, name="Calibri")
            title_cell.alignment = ALIGN_CENTER
            ws.row_dimensions[1].height = 24

            # Style header row (now row 2)
            for col in range(1, max_col+1):
                cell = ws.cell(row=2, column=col)
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
            ws.row_dimensions[2].height = 22

            # Style data rows
            for row in range(3, max_row+2):
                fill = PatternFill("solid", fgColor=C_DARK if row%2==0 else "0A0E1A")
                for col in range(1, max_col+1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.font = Font(color=C_WHITE, size=10, name="Calibri")
                    cell.alignment = ALIGN_LEFT
                    cell.border = THIN_BORDER

            # Freeze header rows
            ws.freeze_panes = ws.cell(row=3, column=1)

            # Auto-width columns (capped)
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 35)

            # Colour-scale on Score/Alpha column if present
            for col in range(1, max_col+1):
                hdr = ws.cell(row=2, column=col).value or ""
                if "Score" in str(hdr) or "Alpha" in str(hdr):
                    col_letter = get_column_letter(col)
                    ws.conditional_formatting.add(
                        f"{col_letter}3:{col_letter}{max_row+1}",
                        ColorScaleRule(start_type="min", start_color="FF4757",
                                       mid_type="percentile", mid_value=50, mid_color="F59E0B",
                                       end_type="max", end_color="10D98D")
                    )
                if "Return" in str(hdr) or "Upside" in str(hdr) or "Momentum" in str(hdr):
                    col_letter = get_column_letter(col)
                    ws.conditional_formatting.add(
                        f"{col_letter}3:{col_letter}{max_row+1}",
                        ColorScaleRule(start_type="min", start_color="FF4757",
                                       mid_type="num", mid_value=0, mid_color="F59E0B",
                                       end_type="max", end_color="10D98D")
                    )

    # Add metadata sheet
    _add_meta_sheet(wb, universe_size, macro_regime)

    # Save styled
    styled_buf = BytesIO()
    wb.save(styled_buf)
    return styled_buf.getvalue()
