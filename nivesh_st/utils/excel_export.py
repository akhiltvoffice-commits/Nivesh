"""
Shared Excel export function used by all pages.
Produces multi-sheet styled workbooks with full fundamentals.
"""
from io import BytesIO
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Colours ────────────────────────────────────────────────────────────────
FILL_HDR  = PatternFill("solid", fgColor="0D111C")
FILL_ODD  = PatternFill("solid", fgColor="111827")
FILL_EVEN = PatternFill("solid", fgColor="0A0E1A")
FILL_GOLD = PatternFill("solid", fgColor="F59E0B")
FILL_NAVY = PatternFill("solid", fgColor="1E3A8A")

FONT_HDR   = Font(bold=True,  color="F1F5F9", size=10, name="Calibri")
FONT_TITLE = Font(bold=True,  color="F59E0B", size=13, name="Calibri")
FONT_BODY  = Font(bold=False, color="CBD5E1", size=10, name="Calibri")
FONT_META  = Font(bold=False, color="8DA4BF", size=9,  name="Calibri")
FONT_KEY   = Font(bold=True,  color="F59E0B", size=9,  name="Calibri")

BORDER = Border(
    left=Side(style="thin",  color="1E2A3A"),
    right=Side(style="thin", color="1E2A3A"),
    bottom=Side(style="thin",color="1E2A3A"),
)
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")

def _style_sheet(ws, title_text: str):
    """Apply NIVESH dark theme to a worksheet with gradient title row."""
    ws.sheet_view.showGridLines = False
    max_col = ws.max_column
    max_row = ws.max_row

    # Insert title row
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.fill = FILL_NAVY; tc.font = FONT_TITLE; tc.alignment = ALIGN_C
    ws.row_dimensions[1].height = 26

    # Header row (now row 2)
    for col in range(1, max_col+1):
        c = ws.cell(row=2, column=col)
        c.fill = FILL_HDR; c.font = FONT_HDR
        c.alignment = ALIGN_C; c.border = BORDER
    ws.row_dimensions[2].height = 20

    # Data rows
    for row in range(3, max_row + 2):
        fill = FILL_ODD if row % 2 == 1 else FILL_EVEN
        for col in range(1, max_col+1):
            c = ws.cell(row=row, column=col)
            if c.value is not None:
                c.fill = fill; c.font = FONT_BODY
                c.alignment = ALIGN_L; c.border = BORDER
    ws.freeze_panes = ws.cell(row=3, column=1)

    # Auto-width — skip MergedCell objects (title row uses merge_cells)
    from openpyxl.cell import MergedCell
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells
                     if not isinstance(c, MergedCell)), default=8)
        first_real = next((c for c in col_cells if not isinstance(c, MergedCell)), None)
        if first_real:
            ws.column_dimensions[first_real.column_letter].width = min(width + 3, 38)

    # Color scale on numeric-looking columns (row 2 = header after title insert)
    from openpyxl.cell import MergedCell as _MC
    for col in range(1, max_col+1):
        c2 = ws.cell(row=2, column=col)
        if isinstance(c2, _MC): continue
        hdr = str(c2.value or "")
        cl  = get_column_letter(col)
        ref = f"{cl}3:{cl}{max_row+1}"
        if any(k in hdr for k in ["Score","Return","Upside","Momentum","Boost"]):
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type="min", start_color="FF4757",
                mid_type="num", mid_value=0, mid_color="F59E0B",
                end_type="max", end_color="10D98D"))
        if "P/E" in hdr or "PEG" in hdr:
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type="min", start_color="10D98D",
                mid_type="percentile", mid_value=50, mid_color="F59E0B",
                end_type="max", end_color="FF4757"))

def build_full_excel(
    main_df: pd.DataFrame,
    universe_df: pd.DataFrame = None,
    macro_regime: dict = None,
    sheet_label: str = "Results",
) -> bytes:
    """
    Multi-sheet Excel:
      Sheet 1 = main_df (whatever the page passes — scored/ranked)
      Sheet 2 = Full Fundamentals (enriched universe_df)
      Sheet 3 = Technicals subset
      Sheet 4 = Glossary & Metadata
    """
    today_str = datetime.now().strftime("%d %b %Y, %H:%M IST")
    n = len(main_df)

    buf = BytesIO()

    # ── Build fundamentals sheet from universe_df ─────────────────────────
    fund_cols = {
        "Symbol":             "Symbol",
        "Name":               "Company",
        "Sector":             "Sector",
        "Price":              "Price (₹)",
        "Market Cap":         "Mkt Cap (₹Cr)",
        "PE":                 "P/E (TTM)",
        "PE (Forward)":       "P/E (Fwd)",
        "PB":                 "P/B",
        "PS":                 "P/S",
        "EV/EBITDA":          "EV/EBITDA",
        "EV/Revenue":         "EV/Revenue",
        "EPS":                "EPS (₹)",
        "PEG":                "PEG Ratio",
        "Graham Number":      "Graham No. (₹)",
        "Graham Premium%":    "Graham Premium %",
        "ROE":                "ROE %",
        "ROA":                "ROA %",
        "ROCE%":              "ROCE %",
        "Net Margin":         "Net Margin %",
        "Operating Margin":   "Oper. Margin %",
        "Gross Margin":       "Gross Margin %",
        "Rev Growth":         "Rev Growth %",
        "EPS Growth":         "EPS Growth %",
        "D/E":                "Debt/Equity",
        "Current Ratio":      "Current Ratio",
        "Interest Coverage":  "Int. Coverage",
        "Payout Ratio":       "Payout Ratio %",
        "Div Yield":          "Div Yield %",
        "Promoter Holding%":  "Promoter Hold %",
        "Institutional Hold%":"Inst. Hold %",
        "Beta":               "Beta",
        "52W%":               "52W Position %",
        "52W High":           "52W High (₹)",
        "52W Low":            "52W Low (₹)",
        "AnalystTarget":      "Analyst Target (₹)",
        "Analyst Upside%":    "Analyst Upside %",
        "Recommendation":     "Consensus",
        "Macro Score":        "Macro Boost",
        "Score":              "Composite Score",
        "Tech":               "Tech Signal",
        "Reasons":            "Score Drivers",
    }

    if universe_df is not None and not universe_df.empty:
        # Merge sector PE if available
        try:
            sector_pe = {}
            for sector, grp in universe_df.groupby("Sector"):
                pe_vals = grp["PE"].dropna()
                pe_vals = pe_vals[(pe_vals > 0) & (pe_vals < 200)]
                if not pe_vals.empty:
                    sector_pe[sector] = round(float(pe_vals.median()), 1)
            universe_df = universe_df.copy()
            universe_df["Sector PE"] = universe_df["Sector"].map(sector_pe)
            universe_df["PE vs Sector"] = universe_df.apply(
                lambda r: round((r["PE"] / r["Sector PE"] - 1)*100, 1)
                if (r.get("PE") and r.get("Sector PE") and r["Sector PE"] > 0) else None, axis=1
            )
        except Exception:
            pass

        fund_cols_ext = dict(fund_cols)
        fund_cols_ext["Sector PE"]    = "Sector Median PE"
        fund_cols_ext["PE vs Sector"] = "PE vs Sector %"

        avail = [c for c in fund_cols_ext if c in universe_df.columns]
        fund_df = universe_df[avail].copy()
        # Format
        for col, fn in [
            ("ROE",lambda x:round(x*100,1)),("ROA",lambda x:round(x*100,1)),
            ("Net Margin",lambda x:round(x*100,1)),("Operating Margin",lambda x:round(x*100,1)),
            ("Gross Margin",lambda x:round(x*100,1)),("Rev Growth",lambda x:round(x*100,1)),
            ("EPS Growth",lambda x:round(x*100,1)),("Div Yield",lambda x:round(x*100,2)),
            ("Payout Ratio",lambda x:round(x*100,1)),
            ("Promoter Holding%",lambda x:round(x*100,1)),
            ("Institutional Hold%",lambda x:round(x*100,1)),
            ("Market Cap",lambda x:round(x/1e7,0) if x else None),
            ("FCF",lambda x:round(x/1e7,0) if x else None),
        ]:
            if col in fund_df.columns:
                fund_df[col] = fund_df[col].map(lambda x,f=fn: f(x) if x else None)
        fund_df = fund_df.rename(columns=fund_cols_ext)
    else:
        fund_df = None

    # ── Write to Excel ────────────────────────────────────────────────────
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        main_df.to_excel(writer, index=False, sheet_name=f"📊 {sheet_label}")
        if fund_df is not None:
            fund_df.to_excel(writer, index=False, sheet_name="💰 Full Fundamentals")
        _write_glossary(writer, macro_regime, today_str, n)

    buf.seek(0)
    wb = load_workbook(buf)
    today_label = datetime.now().strftime("%d %b %Y")
    titles = {
        f"📊 {sheet_label}": f"NIVESH — {sheet_label} | {today_label}",
        "💰 Full Fundamentals": f"Full Fundamental + Valuation Metrics | {today_label}",
    }
    for ws in wb.worksheets:
        if ws.title in titles:
            _style_sheet(ws, titles[ws.title])
        elif "Glossary" in ws.title:
            # Basic dark styling for glossary
            ws.sheet_view.showGridLines = False
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 75
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(color="CBD5E1", size=10, name="Calibri")
                    cell.fill = PatternFill("solid", fgColor="0D111C")
                    if cell.column == 1 and cell.value and not str(cell.value).startswith(" "):
                        cell.font = Font(bold=True, color="F59E0B", size=10, name="Calibri")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def _write_glossary(writer, macro_regime, today_str, n_stocks):
    rows = [
        ["NIVESH Investment Platform", ""],
        ["Download Time", today_str],
        ["Universe", f"{n_stocks} NSE stocks (Nifty50+Next50+Midcap+Smallcap+Extra)"],
        ["", ""],
        ["SCORING FORMULA", ""],
        ["Composite Score",   "Max 120 = Valuation(25) + Quality(30) + Growth(25) + Safety(20) + Technical(10) + Macro(±20)"],
        ["Macro Boost",       "Live sector impact from Brent crude, USD/INR, India VIX, US market, US 10Y rate"],
        ["", ""],
        ["VALUATION METRICS", ""],
        ["P/E Ratio",         "Price ÷ Earnings Per Share (TTM). Lower = cheaper. Industry median more useful than absolute."],
        ["Sector Median PE",  "Median PE of all scored stocks in the same sector. Compare company PE vs this."],
        ["PE vs Sector %",    "Company PE premium (+) or discount (−) vs sector median. Negative = cheaper than peers."],
        ["P/B Ratio",         "Price ÷ Book Value per Share. <1 = trading below asset value."],
        ["P/S Ratio",         "Price ÷ Revenue per Share (TTM). Useful for loss-making/high-growth companies."],
        ["EV/EBITDA",         "Enterprise Value ÷ EBITDA. <10x generally cheap. Compare within sector."],
        ["PEG Ratio",         "P/E ÷ EPS Growth Rate. <0.5 very cheap, 0.5-1 cheap, 1-2 fair, >2 expensive."],
        ["Graham Number",     "√(22.5 × EPS × Book Value/Share). Benjamin Graham intrinsic value estimate."],
        ["Graham Premium %",  "% above (+) or below (−) Graham Number. Negative = below intrinsic value."],
        ["", ""],
        ["QUALITY METRICS", ""],
        ["ROE %",             "Net Income ÷ Equity. >20% excellent, >15% good, <10% weak."],
        ["ROA %",             "Net Income ÷ Total Assets. Higher = more efficient asset use."],
        ["ROCE %",            "Net Income ÷ (Equity + Debt). Best for capital-intensive businesses."],
        ["Net Margin %",      "Net Profit ÷ Revenue. Higher = better pricing power and efficiency."],
        ["Interest Coverage", "EBITDA ÷ Interest Expense. >3x safe, 1.5-3x caution, <1.5x risky."],
        ["Promoter Hold %",   "Insider/promoter holding. >50% = founder-led, aligned. Increasing = positive."],
        ["Institutional Hold%","FII + DII + MF combined holding. Increasing = smart money buying."],
        ["", ""],
        ["TECHNICAL SIGNALS", ""],
        ["RSI (14)",          "Relative Strength Index. <30 oversold (buy zone), >70 overbought (sell zone)."],
        ["52W Position %",    "0% = at 52-week low, 100% = at 52-week high. <20% = near value zone."],
        ["Tech Signal",       "Bullish = price above SMA20+50+200. Bearish = below all three."],
        ["Macro Boost",       "Positive = sector has macro tailwinds. Negative = macro headwinds."],
        ["", ""],
    ]
    if macro_regime:
        rows += [["LIVE MACRO AT DOWNLOAD TIME", ""]]
        raw_map = {
            "crude_price":"Brent Crude ($/bbl)","vix_level":"India VIX",
            "inr_level":"USD/INR","nifty_chg":"Nifty Day Change %",
            "sp_chg":"S&P 500 Day Change %","us10y":"US 10Y Yield %"
        }
        for k, label in raw_map.items():
            v = macro_regime.get(k)
            if v is not None:
                rows.append([label, round(v, 2)])
        active = [k.replace("_"," ").title() for k,v in macro_regime.items() if v is True and isinstance(v,bool)]
        if active:
            rows.append(["Active Flags", " | ".join(active)])

    df = pd.DataFrame(rows, columns=["Key", "Definition"])
    df.to_excel(writer, index=False, sheet_name="📖 Glossary")
